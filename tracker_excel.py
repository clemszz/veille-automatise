"""Intégration de la veille validée dans le classeur Excel de suivi manuel
préexistant (Veille_marché_ENR.xlsx, onglet "Suivi veille") : une ligne par
actualité, rangée dans la bonne colonne thème + acteur/sujet, avec les liens
sources.

Reçoit directement les entrées structurées de l'aperçu (title/summary/
source_line, potentiellement éditées à la main dans la webapp — voir
webapp._parse_edited_entries), sans repasser par le texte formaté ni un
re-parsing : plus robuste (aucune dépendance à la mise en forme exacte du
texte) et évite un aller-retour inutile structuré -> texte -> structuré.
"""
import shutil
import re
from copy import copy
from datetime import date, datetime
from pathlib import Path

import openpyxl

from config import (
    TRACKER_COL_DATE,
    TRACKER_COL_THEMES,
    TRACKER_LIENS_COLUMNS,
    TRACKER_SHEET_NAME,
    TRACKER_THEME_COLUMNS,
    TRACKER_XLSX_PATH,
)
from summarize_mistral import classify_for_tracker

_URL_RE = re.compile(r"https?://\S+")


def extract_links(source_line: str) -> list[str]:
    return _URL_RE.findall(source_line)


LIENS_PLACEHOLDER = "intégrer lien pdf"


def _backup(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = path.with_name(f"{path.stem}.backup-{stamp}{path.suffix}")
    shutil.copy2(path, backup_path)
    return backup_path


def _copy_row_style(ws, template_row: int, target_row: int, max_col: int) -> None:
    """Reprend la mise en forme (police, surlignage, alignement, bordures,
    format de nombre) de template_row sur target_row, colonne par colonne,
    pour que les lignes ajoutées se fondent dans le classeur existant plutôt
    que d'apparaître avec le style par défaut d'openpyxl."""
    for col in range(1, max_col + 1):
        src = ws.cell(row=template_row, column=col)
        dst = ws.cell(row=target_row, column=col)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
    src_height = ws.row_dimensions[template_row].height
    if src_height:
        ws.row_dimensions[target_row].height = src_height


def _find_existing_actor_row(ws, acteur: str, max_row: int) -> int | None:
    """Cherche, parmi les lignes 2..max_row, une ligne déjà associée à cet
    acteur (comparaison insensible à la casse/aux espaces) — pour un acteur
    récurrent (ex. "Nadara"), on la réutilise au lieu d'empiler une nouvelle
    ligne à chaque mention (voir append_entries). Ne s'applique jamais à une
    ligne "DIVERS/ <sujet>" : la comparaison est une égalité stricte avec le
    nom d'acteur, jamais un sous-texte. Ne touche pas à l'historique déjà
    présent avant ce run : une éventuelle répétition déjà existante dans le
    classeur (ex. EDF sur plusieurs lignes) n'est pas fusionnée rétroactivement,
    seules les nouvelles validations réutilisent une ligne trouvée."""
    target = acteur.strip().casefold()
    for r in range(2, max_row + 1):
        val = ws.cell(row=r, column=TRACKER_COL_THEMES).value
        if isinstance(val, str) and val.strip().casefold() == target:
            return r
    return None


def append_entries(
    entries: list[dict], run_date: date, excel_path: Path = TRACKER_XLSX_PATH
) -> list[dict]:
    """entries : dicts avec title/summary/source_line/links/excel_theme/
    acteur/sujet_divers (voir integrate_draft_entries). Sauvegarde le classeur
    avant modification (fichier tenu à la main depuis des années, une
    corruption serait coûteuse à rattraper), puis pour chaque entrée :
    - si un acteur est identifié ET qu'une ligne existante à son nom a la
      case du thème ciblé encore vide (tableau croisé acteur x thème), on
      réutilise cette ligne au lieu d'en créer une nouvelle (voir
      _find_existing_actor_row) — met aussi à jour sa date et complète ses
      colonnes LIENS encore vides sans écraser des liens déjà présents ;
    - sinon (pas d'acteur, pas de ligne existante, ou case déjà occupée par
      une actu précédente — cas volontairement laissé simple pour l'instant)
      on ajoute une nouvelle ligne à la fin, comme avant.
    Retourne les lignes touchées pour affichage/vérification par l'utilisateur."""
    if not excel_path.exists():
        raise FileNotFoundError(f"Classeur introuvable : {excel_path}")

    _backup(excel_path)

    wb = openpyxl.load_workbook(excel_path)
    if TRACKER_SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Onglet '{TRACKER_SHEET_NAME}' introuvable dans {excel_path.name}")
    ws = wb[TRACKER_SHEET_NAME]

    template_row = ws.max_row
    max_col = ws.max_column
    added = []
    next_new_row = template_row + 1
    for e in entries:
        acteur = e.get("acteur")
        sujet = e.get("sujet_divers")
        theme_label = acteur or (f"DIVERS/ {sujet}" if sujet else "DIVERS")
        theme_col = TRACKER_THEME_COLUMNS.get(e["excel_theme"], TRACKER_THEME_COLUMNS["DIVERS"])
        links = e.get("links") or []

        existing_row = _find_existing_actor_row(ws, acteur, ws.max_row) if acteur else None
        if existing_row and not ws.cell(row=existing_row, column=theme_col).value:
            row = existing_row
            ws.cell(row=row, column=theme_col, value=f"{e['title']}\n{e['summary']}")
            ws.cell(row=row, column=TRACKER_COL_DATE, value=run_date)
            empty_link_cols = [c for c in TRACKER_LIENS_COLUMNS if not ws.cell(row=row, column=c).value]
            for link, col in zip(links, empty_link_cols):
                ws.cell(row=row, column=col, value=link)
        else:
            row = next_new_row
            _copy_row_style(ws, template_row, row, max_col)
            ws.cell(row=row, column=TRACKER_COL_THEMES, value=theme_label)
            ws.cell(row=row, column=TRACKER_COL_DATE, value=run_date)
            ws.cell(row=row, column=theme_col, value=f"{e['title']}\n{e['summary']}")
            if links:
                for link, col in zip(links, TRACKER_LIENS_COLUMNS):
                    ws.cell(row=row, column=col, value=link)
            else:
                ws.cell(row=row, column=TRACKER_LIENS_COLUMNS[0], value=LIENS_PLACEHOLDER)
            next_new_row += 1

        added.append({
            "row": row, "theme_label": theme_label, "excel_theme": e["excel_theme"],
            "title": e["title"], "links": links,
        })

    wb.save(excel_path)
    return added


def integrate_draft_entries(
    entries: list[dict], run_date: date, excel_path: Path = TRACKER_XLSX_PATH
) -> list[dict]:
    """Point d'entrée unique utilisé par la webapp (voir webapp._integrate_to_excel) :
    entries = les entrées structurées de l'aperçu déjà validées par
    l'utilisateur (title/summary/source_line). Classe chaque actu (thème
    Excel + acteur/sujet) via Mistral, puis ajoute les lignes au classeur.
    Renvoie la liste vide si `entries` est vide (aperçu validé sans aucune
    actu retenue), au lieu de planter."""
    if not entries:
        return []

    classifications = classify_for_tracker(
        [{"title": e["title"], "summary": e["summary"]} for e in entries]
    )
    combined = [
        {**e, **c, "links": extract_links(e["source_line"])}
        for e, c in zip(entries, classifications)
    ]
    return append_entries(combined, run_date, excel_path)
